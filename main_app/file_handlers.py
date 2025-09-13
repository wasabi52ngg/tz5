# -*- coding: utf-8 -*-
"""
Модульные обработчики файлов для импорта/экспорта контактов
"""
import csv
import io
import logging
from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class BaseFileHandler(ABC):
    """Базовый класс для обработчиков файлов"""
    
    @abstractmethod
    def read_contacts(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Прочитать контакты из файла
        
        Args:
            file_content: Содержимое файла в байтах
            
        Returns:
            Список данных контактов
        """
        pass
    
    @abstractmethod
    def write_contacts(self, contacts_data: List[Dict[str, Any]]) -> bytes:
        """
        Записать контакты в файл
        
        Args:
            contacts_data: Список данных контактов
            
        Returns:
            Содержимое файла в байтах
        """
        pass
    
    @abstractmethod
    def get_file_extension(self) -> str:
        """
        Получить расширение файла
        
        Returns:
            Расширение файла (например, '.csv')
        """
        pass
    
    @abstractmethod
    def get_mime_type(self) -> str:
        """
        Получить MIME-тип файла
        
        Returns:
            MIME-тип файла
        """
        pass


class CSVFileHandler(BaseFileHandler):
    """Обработчик CSV файлов"""
    
    def __init__(self, encoding: str = 'utf-8', delimiter: str = ','):
        self.encoding = encoding
        self.delimiter = delimiter
    
    def read_contacts(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Прочитать контакты из CSV файла
        
        Args:
            file_content: Содержимое файла в байтах
            
        Returns:
            Список данных контактов
        """
        try:
            # Декодируем файл
            text_content = file_content.decode(self.encoding)
            
            # Читаем CSV
            csv_reader = csv.DictReader(io.StringIO(text_content), delimiter=self.delimiter)
            
            contacts = []
            for row in csv_reader:
                # Очищаем данные от лишних пробелов
                cleaned_row = {key.strip(): value.strip() for key, value in row.items()}
                contacts.append(cleaned_row)
            
            return contacts
            
        except Exception as e:
            logger.error(f"Ошибка при чтении CSV файла: {e}")
            raise ValueError(f"Не удалось прочитать CSV файл: {e}")
    
    def write_contacts(self, contacts_data: List[Dict[str, Any]]) -> bytes:
        """
        Записать контакты в CSV файл
        
        Args:
            contacts_data: Список данных контактов
            
        Returns:
            Содержимое файла в байтах
        """
        try:
            if not contacts_data:
                return b''
            
            # Определяем заголовки
            headers = ['имя', 'фамилия', 'отчество', 'номер телефона', 'почта', 'компания']
            
            # Создаем CSV в памяти
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=headers, delimiter=self.delimiter)
            
            # Записываем заголовки
            writer.writeheader()
            
            # Записываем данные
            for contact in contacts_data:
                writer.writerow(contact)
            
            # Возвращаем в байтах
            return output.getvalue().encode(self.encoding)
            
        except Exception as e:
            logger.error(f"Ошибка при записи CSV файла: {e}")
            raise ValueError(f"Не удалось записать CSV файл: {e}")
    
    def get_file_extension(self) -> str:
        return '.csv'
    
    def get_mime_type(self) -> str:
        return 'text/csv'


class XLSXFileHandler(BaseFileHandler):
    """Обработчик XLSX файлов"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите его командой: pip install openpyxl")
    
    def read_contacts(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Прочитать контакты из XLSX файла
        
        Args:
            file_content: Содержимое файла в байтах
            
        Returns:
            Список данных контактов
        """
        try:
            # Создаем временный файл в памяти
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            # Читаем заголовки
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    break
            
            # Читаем данные
            contacts = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Пропускаем пустые строки
                    continue
                
                contact = {}
                for i, value in enumerate(row):
                    if i < len(headers) and value is not None:
                        contact[headers[i]] = str(value).strip()
                
                if contact:  # Добавляем только непустые контакты
                    contacts.append(contact)
            
            return contacts
            
        except Exception as e:
            logger.error(f"Ошибка при чтении XLSX файла: {e}")
            raise ValueError(f"Не удалось прочитать XLSX файл: {e}")
    
    def write_contacts(self, contacts_data: List[Dict[str, Any]]) -> bytes:
        """
        Записать контакты в XLSX файл
        
        Args:
            contacts_data: Список данных контактов
            
        Returns:
            Содержимое файла в байтах
        """
        try:
            # Создаем новую книгу
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Контакты"
            
            # Определяем заголовки
            headers = ['имя', 'фамилия', 'отчество', 'номер телефона', 'почта', 'компания']
            
            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Записываем данные
            for row, contact in enumerate(contacts_data, 2):
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=row, column=col, value=contact.get(header, ''))
            
            # Сохраняем в память
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Ошибка при записи XLSX файла: {e}")
            raise ValueError(f"Не удалось записать XLSX файл: {e}")
    
    def get_file_extension(self) -> str:
        return '.xlsx'
    
    def get_mime_type(self) -> str:
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class FileHandlerFactory:
    """Работа с созданием обработчиков файлов"""
    
    _handlers = {
        '.csv': CSVFileHandler,
        '.xlsx': XLSXFileHandler,
    }
    
    @classmethod
    def get_handler(cls, file_extension: str) -> BaseFileHandler:
        """
        Получить обработчик файла по расширению
        
        Args:
            file_extension: Расширение файла (например, '.csv')
            
        Returns:
            Обработчик файла
            
        Raises:
            ValueError: Если обработчик для данного расширения не найден
        """
        file_extension = file_extension.lower()
        
        if file_extension not in cls._handlers:
            raise ValueError(f"Неподдерживаемый формат файла: {file_extension}")
        
        handler_class = cls._handlers[file_extension]
        
        try:
            return handler_class()
        except ImportError as e:
            raise ImportError(f"Не удалось создать обработчик для {file_extension}: {e}")
    
    @classmethod
    def get_supported_extensions(cls) -> List[str]:
        """
        Получить список поддерживаемых расширений
        
        Returns:
            Список поддерживаемых расширений
        """
        return list(cls._handlers.keys())
    
    @classmethod
    def is_supported(cls, file_extension: str) -> bool:
        """
        Проверить, поддерживается ли расширение файла
        
        Args:
            file_extension: Расширение файла
            
        Returns:
            True, если расширение поддерживается
        """
        return file_extension.lower() in cls._handlers


class FileProcessor:
    """Класс для обработки файлов с контактами"""
    
    def __init__(self, file_extension: str):
        self.handler = FileHandlerFactory.get_handler(file_extension)
    
    def read_contacts(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Прочитать контакты из файла
        
        Args:
            file_content: Содержимое файла в байтах
            
        Returns:
            Список данных контактов
        """
        return self.handler.read_contacts(file_content)
    
    def write_contacts(self, contacts_data: List[Dict[str, Any]]) -> bytes:
        """
        Записать контакты в файл
        
        Args:
            contacts_data: Список данных контактов
            
        Returns:
            Содержимое файла в байтах
        """
        return self.handler.write_contacts(contacts_data)
    
    def get_file_extension(self) -> str:
        return self.handler.get_file_extension()
    
    def get_mime_type(self) -> str:
        return self.handler.get_mime_type()
